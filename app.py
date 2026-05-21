from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

from config import (
    ACTION_LIBRARY,
    APP_SUBTITLE,
    APP_TITLE,
    CHANNEL_ORDER,
    CHANNEL_RULES,
    COLORS,
    EXCEL_FILE,
    EXCLUDE_TITLE_PATTERNS,
    MAX_TOP_OPPORTUNITIES,
    MONTHS,
    MONTH_MAP,
    PAGE_TITLES,
    SIMULATION_DEFAULTS,
    TEXT_HELP,
)

# -----------------------------------------------------------------------------
# Streamlit setup
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title=APP_TITLE,
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
MONTH_SET = set(MONTHS)
CONTROL_TITLE_RE = re.compile(
    "|".join(re.escape(p) for p in EXCLUDE_TITLE_PATTERNS),
    flags=re.IGNORECASE,
)

def normalize_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None

def parse_year_value(value, fallback: int | None = None) -> int | None:
    if value is None:
        return fallback
    text = normalize_text(value)
    if not text:
        return fallback
    match = re.search(r"(\d{4})", text)
    if match:
        return int(match.group(1))
    try:
        return int(float(text))
    except Exception:
        return fallback

def parse_period_num(label):
    text = normalize_text(label)
    if not text:
        return None
    return MONTH_MAP.get(text.upper())

def is_month_period(label) -> bool:
    return parse_period_num(label) is not None

def is_control_title(title: str) -> bool:
    text = normalize_text(title) or ""
    return bool(CONTROL_TITLE_RE.search(text))

def is_aggregate_channel(channel: str) -> bool:
    text = normalize_text(channel) or ""
    upper = text.upper()
    return upper.startswith("TOTAL") or upper.startswith("TOTALES")

def channel_family(channel: str) -> str:
    text = normalize_text(channel)
    if not text:
        return "Otros"
    upper = text.upper()

    for family, tokens in CHANNEL_RULES.items():
        if family == "Otros":
            continue
        if any(token.upper() in upper for token in tokens):
            return family

    if is_aggregate_channel(text):
        return "Totales"

    return "Otros"

def safe_divide(numerator, denominator):
    if denominator in (None, 0) or pd.isna(denominator):
        return np.nan
    return numerator / denominator

def safe_pct_change(current, previous):
    if previous in (None, 0) or pd.isna(previous):
        return np.nan
    return (current - previous) / previous

def fmt_num(value):
    if pd.isna(value):
        return "-"
    return f"{value:,.0f}"

def fmt_mxn(value):
    if pd.isna(value):
        return "-"
    return f"${value:,.0f}"

def fmt_pct(value):
    if pd.isna(value):
        return "-"
    return f"{value:.1%}"

def semaforo(value):
    if pd.isna(value):
        return "⚪", "Sin dato"
    if value >= 8:
        return "🟢", "Alto"
    if value >= 6.5:
        return "🟡", "Medio"
    return "🔴", "Crítico"

def make_header():
    st.markdown(
        f"""
        <div style="background: linear-gradient(135deg, {COLORS['navy']} 0%, {COLORS['info']} 100%);
                    padding: 18px 22px; border-radius: 16px; color: white; margin-bottom: 14px;">
            <div style="font-size: 2rem; font-weight: 800;">{APP_TITLE}</div>
            <div style="opacity: 0.9; margin-top: 4px;">{APP_SUBTITLE}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def make_metric_row(metrics):
    cols = st.columns(len(metrics))
    for col, metric in zip(cols, metrics):
        with col:
            st.metric(**metric)

def title_rows(ws):
    found = []
    seen = set()
    for r in range(1, ws.max_row + 1):
        title = None
        title_col = None
        for c in (1, 2):
            v = normalize_text(ws.cell(r, c).value)
            if v and v.startswith("Aforo e Ingreso -"):
                title = v.replace("Aforo e Ingreso -", "").strip()
                title_col = c
                break
        if title and r not in seen:
            found.append((r, title_col, title))
            seen.add(r)
    return found

def header_row_after(ws, title_row, lookahead=6):
    for r in range(title_row + 1, min(ws.max_row, title_row + lookahead) + 1):
        for c in range(1, min(ws.max_column, 4) + 1):
            if normalize_text(ws.cell(r, c).value) == "FECHA":
                return r, c
    return None, None

def collect_channel_labels(ws, header_row, fecha_col):
    labels = []
    col = fecha_col + 2
    while col <= ws.max_column:
        label = normalize_text(ws.cell(header_row - 1, col).value) if header_row - 1 >= 1 else None
        if not label:
            break
        labels.append(label)
        col += 2
    return labels

def parse_hoja2(ws):
    if ws is None:
        return pd.DataFrame()

    records = []
    for row in range(4, ws.max_row + 1):
        periodo = normalize_text(ws.cell(row, 1).value)
        if not periodo:
            continue

        values = [ws.cell(row, c).value for c in range(2, 8)]
        if all(v is None for v in values):
            continue

        records.append(
            {
                "periodo": periodo,
                "mis_en_sus_aforo": ws.cell(row, 2).value,
                "mis_en_sus_ingreso": ws.cell(row, 3).value,
                "mis_en_mis_aforo": ws.cell(row, 4).value,
                "mis_en_mis_ingreso": ws.cell(row, 5).value,
                "sus_en_mis_aforo": ws.cell(row, 6).value,
                "sus_en_mis_ingreso": ws.cell(row, 7).value,
            }
        )
    return pd.DataFrame(records)

def parse_workbook(file_path):
    wb = load_workbook(file_path, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Hoja2":
            continue

        ws = wb[sheet_name]
        sheet_year = parse_year_value(sheet_name)

        titles = title_rows(ws)

        for idx, (tr, _, title) in enumerate(titles):
            header_row, fecha_col = header_row_after(ws, tr)
            if header_row is None:
                continue

            next_title_row = titles[idx + 1][0] if idx + 1 < len(titles) else ws.max_row + 1
            channel_labels = collect_channel_labels(ws, header_row, fecha_col)

            if len(channel_labels) < 1:
                continue

            block_kind = "control" if is_control_title(title) else "operational"

            for row in range(header_row + 2, next_title_row):
                period = normalize_text(ws.cell(row, fecha_col).value)
                if not period:
                    continue

                period_upper = period.upper()
                if period_upper not in MONTH_SET and period_upper not in {"TOTALES", "PROMEDIO"}:
                    continue

                row_year = parse_year_value(ws.cell(row, fecha_col + 1).value, fallback=sheet_year)

                for channel_idx, channel in enumerate(channel_labels):
                    aforo_col = fecha_col + 2 + channel_idx * 2
                    ingreso_col = aforo_col + 1

                    if aforo_col > ws.max_column:
                        break

                    aforo = ws.cell(row, aforo_col).value
                    ingreso = ws.cell(row, ingreso_col).value if ingreso_col <= ws.max_column else None

                    if aforo is None and ingreso is None:
                        continue

                    records.append(
                        {
                            "sheet": sheet_name,
                            "year": row_year,
                            "concession": title,
                            "title_row": tr,
                            "header_row": header_row,
                            "fecha_col": fecha_col,
                            "period": period_upper,
                            "month_num": parse_period_num(period_upper),
                            "period_type": "month" if is_month_period(period_upper) else "summary",
                            "channel": channel,
                            "family": channel_family(channel),
                            "is_aggregate_channel": is_aggregate_channel(channel),
                            "aforo": float(aforo) if aforo is not None else 0.0,
                            "ingreso": float(ingreso) if ingreso is not None else 0.0,
                            "block_kind": block_kind,
                        }
                    )

    raw_df = pd.DataFrame(records)

    if not raw_df.empty:
        raw_df["concession"] = raw_df["concession"].astype(str).str.strip()
        raw_df["channel"] = raw_df["channel"].astype(str).str.strip()
        raw_df["family"] = raw_df["family"].astype(str).str.strip()
        raw_df["block_kind"] = raw_df["block_kind"].astype(str).str.strip()
        raw_df["year"] = pd.to_numeric(raw_df["year"], errors="coerce").astype("Int64")
        raw_df["month_num"] = pd.to_numeric(raw_df["month_num"], errors="coerce").astype("Int64")

    hoja2_df = parse_hoja2(wb["Hoja2"]) if "Hoja2" in wb.sheetnames else pd.DataFrame()
    return raw_df, hoja2_df

@st.cache_data(show_spinner=False)
def load_data(file_path):
    return parse_workbook(file_path)

def prepare_analysis(raw_df, scope_mode, year_choice, concession_choice):
    base = raw_df.copy()

    if scope_mode == TEXT_HELP["scope_operational"]:
        base = base[base["block_kind"] == "operational"].copy()

    base = base[base["period_type"] == "month"].copy()
    base = base[~base["is_aggregate_channel"]].copy()

    if year_choice != "Todos":
        base = base[base["year"] == int(year_choice)].copy()

    if concession_choice != "Todas":
        base = base[base["concession"] == concession_choice].copy()

    base = base.dropna(subset=["year", "month_num"])

    if base.empty:
        return base, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    base["year"] = base["year"].astype(int)
    base["month_num"] = base["month_num"].astype(int)

    monthly = (
        base.groupby(["year", "month_num"], as_index=False)
        .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
        .sort_values(["year", "month_num"])
    )

    concession_summary = (
        base.groupby(["year", "concession"], as_index=False)
        .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
        .sort_values(["year", "aforo"], ascending=[True, False])
    )
    concession_summary["rpc"] = concession_summary["ingreso"] / concession_summary["aforo"].replace(0, np.nan)

    family_mix = (
        base.groupby(["year", "month_num", "family"], as_index=False)
        .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
        .sort_values(["year", "month_num", "family"])
    )

    annual = (
        base.groupby(["year"], as_index=False)
        .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
        .sort_values("year")
    )
    annual["rpc"] = annual["ingreso"] / annual["aforo"].replace(0, np.nan)
    annual["aforo_yoy"] = annual["aforo"].pct_change()
    annual["ingreso_yoy"] = annual["ingreso"].pct_change()

    return base, monthly, concession_summary, family_mix, annual

def build_concession_diagnostics(df_months, selected_year):
    year_df = df_months[df_months["year"] == selected_year].copy()
    if year_df.empty:
        return pd.DataFrame()

    summary = (
        year_df.groupby("concession", as_index=False)
        .agg(
            aforo=("aforo", "sum"),
            ingreso=("ingreso", "sum"),
            aforo_std=("aforo", "std"),
            aforo_mean=("aforo", "mean"),
        )
    )
    summary["rpc"] = summary["ingreso"] / summary["aforo"].replace(0, np.nan)
    summary["volatility"] = summary["aforo_std"] / summary["aforo_mean"].replace(0, np.nan)

    prev = (
        df_months[df_months["year"] == selected_year - 1]
        .groupby("concession", as_index=False)
        .agg(
            aforo_prev=("aforo", "sum"),
            ingreso_prev=("ingreso", "sum"),
        )
    )
    diag = summary.merge(prev, on="concession", how="left")
    diag["aforo_yoy"] = diag.apply(lambda r: safe_pct_change(r["aforo"], r["aforo_prev"]), axis=1)
    diag["ingreso_yoy"] = diag.apply(lambda r: safe_pct_change(r["ingreso"], r["ingreso_prev"]), axis=1)

    family = (
        year_df.groupby(["concession", "family"], as_index=False)
        .agg(aforo=("aforo", "sum"))
    )
    family_pivot = (
        family.pivot(index="concession", columns="family", values="aforo")
        .fillna(0)
    )
    for fam in CHANNEL_ORDER:
        if fam not in family_pivot.columns:
            family_pivot[fam] = 0
    family_pivot = family_pivot[CHANNEL_ORDER]
    family_pivot["aforo_total_family"] = family_pivot.sum(axis=1)
    family_pivot["televia_share"] = family_pivot["TeleVía"] / family_pivot["aforo_total_family"].replace(0, np.nan)
    family_pivot["pase_share"] = family_pivot["PASE"] / family_pivot["aforo_total_family"].replace(0, np.nan)
    family_pivot["premium_share"] = (
        family_pivot[["TeleVía", "CAPUFE/Exentos", "PINFRA", "SITEL", "EASYTRIP"]]
        .sum(axis=1)
        / family_pivot["aforo_total_family"].replace(0, np.nan)
    )
    diag = diag.merge(
        family_pivot[["televia_share", "pase_share", "premium_share"]],
        left_on="concession",
        right_index=True,
        how="left",
    )
    return diag.sort_values("aforo", ascending=False)

def build_heatmap(df_months, selected_year):
    year_df = df_months[df_months["year"] == selected_year].copy()
    if year_df.empty:
        return pd.DataFrame()

    pivot = (
        year_df.groupby(["concession", "family"], as_index=False)
        .agg(aforo=("aforo", "sum"))
        .pivot(index="concession", columns="family", values="aforo")
        .fillna(0)
    )

    cols = [c for c in CHANNEL_ORDER if c in pivot.columns]
    pivot = pivot.reindex(columns=cols, fill_value=0)
    pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]
    return pivot

def build_opportunities(diag_df):
    if diag_df.empty:
        return diag_df

    work = diag_df.copy()
    required_cols = ["aforo", "rpc", "aforo_yoy", "volatility", "pase_share", "televia_share"]
    for col in required_cols:
        if col not in work.columns:
            work[col] = np.nan

    work["max_aforo"] = work["aforo"].max()
    work["median_rpc"] = work["rpc"].median()

    work["volume_component"] = work["aforo"] / work["max_aforo"].replace(0, np.nan)
    work["rpc_gap"] = ((work["median_rpc"] - work["rpc"]) / work["median_rpc"]).clip(lower=0)
    work["drop_component"] = (-work["aforo_yoy"]).clip(lower=0)
    work["vol_component"] = work["volatility"].fillna(0).clip(upper=1.5)

    work["priority_score"] = (
        0.35 * work["volume_component"].fillna(0)
        + 0.25 * work["rpc_gap"].fillna(0)
        + 0.25 * work["drop_component"].fillna(0)
        + 0.15 * work["vol_component"].fillna(0)
    )

    def recommend(row):
        if pd.notna(row["aforo_yoy"]) and row["aforo_yoy"] < 0 and row["aforo"] >= work["aforo"].median():
            return ACTION_LIBRARY["volume_recovery"]["label"], "Base relevante con caída vs año previo."
        if pd.notna(row["rpc"]) and row["rpc"] < row["median_rpc"]:
            return ACTION_LIBRARY["monetization"]["label"], "Ingreso por cruce por debajo del benchmark."
        if pd.notna(row["volatility"]) and row["volatility"] > work["volatility"].median():
            return ACTION_LIBRARY["stabilization"]["label"], "Alta variabilidad mensual."
        if pd.notna(row["pase_share"]) and row["pase_share"] > 0.20:
            return ACTION_LIBRARY["channel_capture"]["label"], "Participación PASE relevante; oportunidad de captura."
        if pd.notna(row["televia_share"]) and row["televia_share"] < 0.25 and row["aforo"] >= work["aforo"].median():
            return ACTION_LIBRARY["premium_mix"]["label"], "Mix TeleVía bajo para una base relevante."
        return "Seguimiento", "Mantener monitoreo y buscar mejoras incrementales."

    labels = work.apply(recommend, axis=1, result_type="expand")
    work["accion"] = labels[0]
    work["motivo"] = labels[1]
    return work.sort_values("priority_score", ascending=False)

def make_bar_config(fig, height=420):
    fig.update_layout(
        height=height,
        plot_bgcolor="white",
        paper_bgcolor="white",
        margin=dict(l=10, r=10, t=20, b=10),
        legend_title_text="",
        font=dict(size=11),
    )
    return fig

def render_table(df, cols=None, formatters=None, height=320):
    if df.empty:
        st.info("Sin datos para mostrar.")
        return
    out = df.copy()
    if cols is not None:
        out = out[cols]
    if formatters:
        st.dataframe(out.style.format(formatters), use_container_width=True, height=height, hide_index=True)
    else:
        st.dataframe(out, use_container_width=True, height=height, hide_index=True)

# -----------------------------------------------------------------------------
# Load data
# -----------------------------------------------------------------------------
excel_path = Path(EXCEL_FILE)
if not excel_path.exists():
    st.error(f"No encuentro el archivo {EXCEL_FILE} en el repositorio.")
    xlsx_files = [p.name for p in Path(".").glob("*.xlsx")]
    if xlsx_files:
        st.info("Archivos XLSX visibles: " + ", ".join(xlsx_files))
    st.stop()

with st.spinner("Cargando y estructurando el Excel..."):
    raw_df, hoja2_df = load_data(str(excel_path))

if raw_df.empty:
    st.error("No se pudo estructurar el archivo. Revisa que el libro conserve el formato de aforos.")
    st.stop()

all_years = sorted(raw_df["year"].dropna().astype(int).unique().tolist())
all_concessions = sorted(
    raw_df.loc[~raw_df["concession"].map(is_control_title), "concession"]
    .dropna()
    .astype(str)
    .unique()
    .tolist()
)

# -----------------------------------------------------------------------------
# Sidebar
# -----------------------------------------------------------------------------
with st.sidebar:
    st.markdown("### Filtros globales")
    scope_mode = st.radio(
        "Alcance",
        [TEXT_HELP["scope_operational"], TEXT_HELP["scope_all"]],
        index=0,
    )
    year_choice = st.selectbox("Año", ["Todos"] + all_years, index=0)
    concession_choice = st.selectbox("Concesión", ["Todas"] + all_concessions, index=0)

    st.markdown("---")
    st.markdown("### Supuestos de simulación")
    sim_volume_growth = st.slider(
        "Crecimiento de volumen (%)",
        0.0, 20.0,
        float(SIMULATION_DEFAULTS["volume_growth_pct"]),
        0.1,
    )
    sim_share_capture = st.slider(
        "Captura de share (pp)",
        0.0, 10.0,
        float(SIMULATION_DEFAULTS["share_capture_pp"]),
        0.1,
    )
    sim_rpc_uplift = st.slider(
        "Mejora ingreso/cruce (%)",
        0.0, 20.0,
        float(SIMULATION_DEFAULTS["rpc_uplift_pct"]),
        0.1,
    )
    # Nuevo control para captura de SUS EN MIS hacia TeleVía
    sim_sus_capture = st.slider(
        "Captura de SUS EN MIS hacia TeleVía (%)",
        0.0, 100.0,
        0.0,
        1.0,
        help="Porcentaje de vehículos de otros operadores (SUS EN MIS) que se convierten a TeleVía."
    )

# -----------------------------------------------------------------------------
# Filtering and derivations
# -----------------------------------------------------------------------------
analysis_df, monthly_df, concession_summary_df, family_mix_df, annual_df = prepare_analysis(
    raw_df=raw_df,
    scope_mode=scope_mode,
    year_choice=year_choice,
    concession_choice=concession_choice,
)

if analysis_df.empty:
    st.error("No hay datos en el alcance actual. Ajusta los filtros laterales.")
    st.stop()

# Serie anual sin filtro por año / concesión para la pestaña de seguimiento.
scope_for_annual = raw_df.copy()
if scope_mode == TEXT_HELP["scope_operational"]:
    scope_for_annual = scope_for_annual[scope_for_annual["block_kind"] == "operational"].copy()
scope_for_annual = scope_for_annual[
    (scope_for_annual["period_type"] == "month")
    & (~scope_for_annual["is_aggregate_channel"])
].copy()
if concession_choice != "Todas":
    scope_for_annual = scope_for_annual[scope_for_annual["concession"] == concession_choice].copy()
scope_for_annual = scope_for_annual.dropna(subset=["year", "month_num"])
scope_for_annual["year"] = scope_for_annual["year"].astype(int)
scope_for_annual["month_num"] = scope_for_annual["month_num"].astype(int)

annual_all = (
    scope_for_annual.groupby("year", as_index=False)
    .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
    .sort_values("year")
)
annual_all["rpc"] = annual_all["ingreso"] / annual_all["aforo"].replace(0, np.nan)
annual_all["aforo_yoy"] = annual_all["aforo"].pct_change()
annual_all["ingreso_yoy"] = annual_all["ingreso"].pct_change()

selected_year = int(year_choice) if year_choice != "Todos" else int(analysis_df["year"].max())
year_df = analysis_df[analysis_df["year"] == selected_year].copy()
previous_year_df = analysis_df[analysis_df["year"] == selected_year - 1].copy()

monthly_year = (
    year_df.groupby("month_num", as_index=False)
    .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
    .sort_values("month_num")
)

prev_monthly_year = (
    previous_year_df.groupby("month_num", as_index=False)
    .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
    .sort_values("month_num")
) if not previous_year_df.empty else pd.DataFrame(columns=["month_num", "aforo", "ingreso"])

focus_month = int(monthly_year["month_num"].max()) if not monthly_year.empty else 12
current_ytd_df = year_df[year_df["month_num"] <= focus_month].copy()
previous_ytd_df = previous_year_df[previous_year_df["month_num"] <= focus_month].copy()

portfolio_aforo = current_ytd_df["aforo"].sum()
portfolio_ingreso = current_ytd_df["ingreso"].sum()
portfolio_rpc = safe_divide(portfolio_ingreso, portfolio_aforo)

prev_aforo = previous_ytd_df["aforo"].sum() if not previous_ytd_df.empty else np.nan
prev_ingreso = previous_ytd_df["ingreso"].sum() if not previous_ytd_df.empty else np.nan

aforo_yoy = safe_pct_change(portfolio_aforo, prev_aforo)
ingreso_yoy = safe_pct_change(portfolio_ingreso, prev_ingreso)

# Concession summary for selected year / YTD
conc_year = (
    current_ytd_df.groupby("concession", as_index=False)
    .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
    .sort_values("aforo", ascending=False)
)
conc_year["rpc"] = conc_year["ingreso"] / conc_year["aforo"].replace(0, np.nan)

total_aforo = conc_year["aforo"].sum()
conc_year["share_aforo"] = conc_year["aforo"] / total_aforo if total_aforo != 0 else np.nan

conc_year["aforo_rank"] = np.arange(1, len(conc_year) + 1)

prev_conc = (
    previous_ytd_df.groupby("concession", as_index=False)
    .agg(aforo_prev=("aforo", "sum"), ingreso_prev=("ingreso", "sum"))
)

conc_diag = conc_year.merge(prev_conc, on="concession", how="left")
conc_diag["aforo_yoy"] = conc_diag.apply(lambda r: safe_pct_change(r["aforo"], r["aforo_prev"]), axis=1)
conc_diag["ingreso_yoy"] = conc_diag.apply(lambda r: safe_pct_change(r["ingreso"], r["ingreso_prev"]), axis=1)

concession_monthly_vol = (
    current_ytd_df.groupby("concession")["aforo"]
    .agg(["mean", "std"])
    .reset_index()
)
concession_monthly_vol["volatility"] = concession_monthly_vol["std"] / concession_monthly_vol["mean"].replace(0, np.nan)
conc_diag = conc_diag.merge(concession_monthly_vol[["concession", "volatility"]], on="concession", how="left")

family_year = (
    current_ytd_df.groupby("family", as_index=False)
    .agg(aforo=("aforo", "sum"), ingreso=("ingreso", "sum"))
    .sort_values("aforo", ascending=False)
)
total_family_aforo = family_year["aforo"].sum()
family_year["share"] = family_year["aforo"] / total_family_aforo if total_family_aforo != 0 else np.nan

heatmap_df = build_heatmap(analysis_df, selected_year)

# Preparar DataFrame para oportunidades con las columnas de familias
family_pivot = (
    current_ytd_df.groupby(["concession", "family"], as_index=False)
    .agg(aforo=("aforo", "sum"))
    .pivot(index="concession", columns="family", values="aforo")
    .fillna(0)
)
for fam in CHANNEL_ORDER:
    if fam not in family_pivot.columns:
        family_pivot[fam] = 0
family_pivot = family_pivot[CHANNEL_ORDER]
family_pivot["televia_share"] = family_pivot["TeleVía"] / family_pivot.sum(axis=1).replace(0, np.nan)
family_pivot["pase_share"] = family_pivot["PASE"] / family_pivot.sum(axis=1).replace(0, np.nan)

opportunities_df = build_opportunities(
    conc_diag.merge(
        family_pivot,
        left_on="concession",
        right_index=True,
        how="left",
    )
)

# normalize family share columns if present
for fam in ["TeleVía", "PASE", "CAPUFE/Exentos", "PINFRA", "SITEL", "EASYTRIP", "Otros", "Totales"]:
    if fam not in opportunities_df.columns:
        opportunities_df[fam] = 0

opportunities_df["family_total"] = opportunities_df[["TeleVía", "PASE", "CAPUFE/Exentos", "PINFRA", "SITEL", "EASYTRIP", "Otros"]].sum(axis=1)
opportunities_df["televia_share"] = opportunities_df["TeleVía"] / opportunities_df["family_total"].replace(0, np.nan)
opportunities_df["pase_share"] = opportunities_df["PASE"] / opportunities_df["family_total"].replace(0, np.nan)

# Top narrative insights
best_concession = conc_year.iloc[0]["concession"] if not conc_year.empty else "-"
worst_concession = conc_year.iloc[-1]["concession"] if not conc_year.empty else "-"
best_rpc_concession = conc_year.sort_values("rpc", ascending=False).iloc[0]["concession"] if not conc_year.empty else "-"
worst_rpc_concession = conc_year.sort_values("rpc", ascending=True).iloc[0]["concession"] if not conc_year.empty else "-"
best_family = family_year.iloc[0]["family"] if not family_year.empty else "-"
worst_family = family_year.iloc[-1]["family"] if not family_year.empty else "-"

# -----------------------------------------------------------------------------
# Header
# -----------------------------------------------------------------------------
make_header()
st.caption(
    f"Archivo fuente: {EXCEL_FILE} · Rango analizado: {min(all_years)}–{max(all_years)} · "
    f"Alcance activo: {scope_mode} · Corte: {selected_year} YTD hasta {focus_month}"
)

# -----------------------------------------------------------------------------
# Tabs
# -----------------------------------------------------------------------------
tab_home, tab_diag, tab_sim, tab_plan, tab_follow, tab_interop = st.tabs([
    PAGE_TITLES["home"],
    PAGE_TITLES["diagnostico"],
    PAGE_TITLES["simulacion"],
    PAGE_TITLES["plan"],
    PAGE_TITLES["seguimiento"],
    "Interoperabilidad y Captura",  # Nueva pestaña
])

# -----------------------------------------------------------------------------
# 1) Executive Summary
# -----------------------------------------------------------------------------
with tab_home:
    st.subheader("Executive Summary")

    make_metric_row([
        {"label": "Aforo total", "value": fmt_num(portfolio_aforo), "delta": fmt_pct(aforo_yoy)},
        {"label": "Ingreso total", "value": fmt_mxn(portfolio_ingreso), "delta": fmt_pct(ingreso_yoy)},
        {"label": "Ingreso / cruce", "value": f"{portfolio_rpc:.2f}" if pd.notna(portfolio_rpc) else "-", "delta": f"{selected_year} YTD"},
        {"label": "Mejor concesión", "value": best_concession, "delta": fmt_num(conc_year.iloc[0]["aforo"]) if not conc_year.empty else "-"},
    ])

    insight_cols = st.columns(4)
    with insight_cols[0]:
        st.info(f"**Mayor volumen:** {best_concession}")
    with insight_cols[1]:
        st.success(f"**RPC líder:** {best_rpc_concession}")
    with insight_cols[2]:
        st.error(f"**RPC más bajo:** {worst_rpc_concession}")
    with insight_cols[3]:
        st.warning(f"**Mejor familia:** {best_family}")

    st.markdown("#### Lectura ejecutiva")
    left, right = st.columns([1.2, 0.8])

    with left:
        st.markdown("##### Tendencia mensual de aforo y ingreso")
        fig = go.Figure()
        if not prev_monthly_year.empty:
            fig.add_trace(
                go.Scatter(
                    x=prev_monthly_year["month_num"],
                    y=prev_monthly_year["aforo"],
                    mode="lines+markers",
                    name=str(selected_year - 1),
                    line=dict(color=COLORS["muted"], width=3),
                )
            )
        fig.add_trace(
            go.Scatter(
                x=monthly_year["month_num"],
                y=monthly_year["aforo"],
                mode="lines+markers",
                name=str(selected_year),
                line=dict(color=COLORS["televia"], width=4),
            )
        )
        fig.add_trace(
            go.Scatter(
                x=monthly_year["month_num"],
                y=monthly_year["ingreso"] / max(monthly_year["ingreso"].max(), 1) * max(monthly_year["aforo"].max(), 1),
                mode="lines+markers",
                name="Ingreso (escala ajustada)",
                line=dict(color=COLORS["info"], width=3, dash="dot"),
                visible="legendonly",
            )
        )
        fig.update_layout(
            xaxis=dict(
                title="",
                tickmode="array",
                tickvals=list(range(1, 13)),
                ticktext=MONTHS,
            ),
            yaxis=dict(title="Aforo", gridcolor=COLORS["grid"]),
            legend_title_text="",
        )
        st.plotly_chart(make_bar_config(fig, height=380), use_container_width=True)

    with right:
        st.markdown("##### Mix de familias")
        family_mix = family_year[family_year["family"] != "Totales"].copy()
        if not family_mix.empty:
            fig = px.pie(
                family_mix,
                names="family",
                values="aforo",
                hole=0.5,
                color_discrete_sequence=[COLORS["televia"], COLORS["pase"], COLORS["info"], COLORS["warning"], COLORS["purple"], COLORS["danger"], COLORS["muted"]],
            )
            fig.update_layout(margin=dict(l=10, r=10, t=20, b=10), height=380)
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("##### Top concesiones por volumen")
    top10 = conc_year.head(10).copy()
    if not top10.empty:
        top10["rpc"] = top10["rpc"].round(2)
        st.dataframe(
            top10[["concession", "aforo", "ingreso", "rpc", "share_aforo"]].rename(
                columns={
                    "concession": "Concesión",
                    "aforo": "Aforo",
                    "ingreso": "Ingreso",
                    "rpc": "Ingreso / cruce",
                    "share_aforo": "Share aforo",
                }
            ).style.format({
                "Aforo": "{:,.0f}",
                "Ingreso": "${:,.0f}",
                "Ingreso / cruce": "{:,.2f}",
                "Share aforo": "{:.1%}",
            }),
            use_container_width=True,
            height=320,
        )

    st.markdown("##### Señales clave")
    signals = [
        f"**Familia líder:** {best_family}",
        f"**Familia rezagada:** {worst_family}",
        f"**Top concesión:** {best_concession}",
        f"**Concesión en atención:** {worst_concession}",
    ]
    st.info(" · ".join(signals))

# -----------------------------------------------------------------------------
# 2) Diagnostic
# -----------------------------------------------------------------------------
with tab_diag:
    st.subheader("Diagnóstico")

    diag_year = st.selectbox("Año para diagnóstico", all_years, index=len(all_years) - 1, key="diag_year")
    diag_year_df = analysis_df[analysis_df["year"] == int(diag_year)].copy()

    if diag_year_df.empty:
        st.info("No hay datos para el año seleccionado.")
    else:
        diag = build_concession_diagnostics(analysis_df, int(diag_year))
        if diag.empty:
            st.info("No hay suficientes datos para generar diagnóstico.")
        else:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Concesiones", f"{diag['concession'].nunique():,}")
            c2.metric("Aforo máximo", fmt_num(diag["aforo"].max()))
            c3.metric("Aforo mínimo", fmt_num(diag["aforo"].min()))
            c4.metric("RPC mediano", f"{diag['rpc'].median():.2f}")

            cols = st.columns([1.05, 0.95])
            with cols[0]:
                st.markdown("##### Ranking de concesiones")
                plot_df = diag.sort_values("aforo", ascending=False).head(15).copy()
                fig = px.bar(
                    plot_df,
                    x="concession",
                    y="aforo",
                    color="rpc",
                    color_continuous_scale="RdYlGn",
                    text_auto=".0f",
                    hover_data={
                        "ingreso": ":,.0f",
                        "aforo_yoy": ":.1%",
                        "volatility": ":.2f",
                    },
                )
                fig.update_layout(
                    xaxis_title="",
                    yaxis_title="Aforo",
                    coloraxis_colorbar=dict(title="RPC"),
                )
                st.plotly_chart(make_bar_config(fig, height=420), use_container_width=True)

            with cols[1]:
                st.markdown("##### Variación vs año anterior")
                yoy_df = diag.sort_values("aforo_yoy", ascending=True).tail(15).copy()
                fig = px.bar(
                    yoy_df,
                    x="aforo_yoy",
                    y="concession",
                    orientation="h",
                    color="aforo_yoy",
                    color_continuous_scale="RdYlGn",
                    text_auto=".1%",
                )
                fig.update_layout(
                    xaxis_title="% YoY aforo",
                    yaxis_title="",
                    coloraxis_colorbar=dict(title="% YoY"),
                )
                st.plotly_chart(make_bar_config(fig, height=420), use_container_width=True)

            st.markdown("##### Heatmap de mix por concesión")
            heatmap = build_heatmap(analysis_df, int(diag_year))
            if not heatmap.empty:
                heatmap_pct = heatmap.div(heatmap.sum(axis=1).replace(0, np.nan), axis=0).fillna(0)
                fig = go.Figure(
                    data=go.Heatmap(
                        z=heatmap_pct.values,
                        x=heatmap_pct.columns.tolist(),
                        y=heatmap_pct.index.tolist(),
                        colorscale="RdYlGn",
                        zmin=0,
                        zmax=float(np.nanmax(heatmap_pct.values)) if np.isfinite(np.nanmax(heatmap_pct.values)) else 1,
                        colorbar=dict(title="Mix"),
                        hovertemplate="Concesión=%{y}<br>Familia=%{x}<br>Mix=%{z:.1%}<extra></extra>",
                    )
                )
                fig.update_layout(height=max(420, 26 * len(heatmap_pct) + 140), plot_bgcolor="white", paper_bgcolor="white")
                st.plotly_chart(fig, use_container_width=True)

            st.markdown("##### Tabla diagnóstica")
            diag_table = diag[
                ["concession", "aforo", "ingreso", "rpc", "aforo_yoy", "ingreso_yoy", "volatility", "televia_share", "pase_share", "premium_share"]
            ].copy()
            diag_table.columns = [
                "Concesión",
                "Aforo",
                "Ingreso",
                "Ingreso/cruce",
                "YoY aforo",
                "YoY ingreso",
                "Volatilidad",
                "TeleVía %",
                "PASE %",
                "Premium %",
            ]
            st.dataframe(
                diag_table.style.format({
                    "Aforo": "{:,.0f}",
                    "Ingreso": "${:,.0f}",
                    "Ingreso/cruce": "{:,.2f}",
                    "YoY aforo": "{:.1%}",
                    "YoY ingreso": "{:.1%}",
                    "Volatilidad": "{:.2f}",
                    "TeleVía %": "{:.1%}",
                    "PASE %": "{:.1%}",
                    "Premium %": "{:.1%}",
                }),
                use_container_width=True,
                height=340,
            )

    # ----- Alertas de fuga (a nivel portafolio anual) -----
    st.markdown("#### Alertas de fuga de tráfico propio")
    if not hoja2_df.empty:
        # Calcular ratio de fuga = MIS EN SUS / MIS EN MIS
        hoja2_df["fuga_ratio"] = hoja2_df["mis_en_sus_aforo"] / hoja2_df["mis_en_mis_aforo"].replace(0, np.nan)
        # Mostrar años donde ratio > 0.3
        fuga_alerta = hoja2_df[hoja2_df["fuga_ratio"] > 0.3].copy()
        if not fuga_alerta.empty:
            st.warning("Se detectaron años con alta fuga de tráfico propio (ratio > 0.3).")
            fuga_table = fuga_alerta[["periodo", "mis_en_mis_aforo", "mis_en_sus_aforo", "fuga_ratio"]].copy()
            fuga_table.columns = ["Año", "Aforo MIS EN MIS", "Aforo MIS EN SUS", "Ratio fuga"]
            st.dataframe(
                fuga_table.style.format({
                    "Aforo MIS EN MIS": "{:,.0f}",
                    "Aforo MIS EN SUS": "{:,.0f}",
                    "Ratio fuga": "{:.2%}",
                }),
                use_container_width=True,
                height=200,
                hide_index=True,
            )
            st.info("**Sugerencia:** Implementar campañas de retención y mejorar beneficios de TeleVía para reducir la fuga.")
        else:
            st.success("No se detectaron años con fuga significativa (ratio <= 0.3).")
    else:
        st.info("No hay datos de interoperabilidad (Hoja2) para calcular fuga.")

# -----------------------------------------------------------------------------
# 3) Simulation (mejorada con captura de SUS EN MIS)
# -----------------------------------------------------------------------------
with tab_sim:
    st.subheader("Simulación")

    sim_target = st.selectbox("Base de simulación", ["Portafolio completo"] + all_concessions, index=0, key="sim_target")

    if sim_target == "Portafolio completo":
        sim_base = current_ytd_df.copy()
    else:
        sim_base = current_ytd_df[current_ytd_df["concession"] == sim_target].copy()

    if sim_base.empty:
        st.info("No hay datos suficientes para simular esa base.")
    else:
        base_aforo = sim_base["aforo"].sum()
        base_ingreso = sim_base["ingreso"].sum()
        base_rpc = safe_divide(base_ingreso, base_aforo)

        # Obtener volumen de SUS EN MIS para el año seleccionado (desde hoja2_df)
        sus_volumen = 0
        if not hoja2_df.empty and sim_target == "Portafolio completo":
            año_actual = selected_year if year_choice != "Todos" else annual_all["year"].max()
            año_row = hoja2_df[hoja2_df["periodo"] == str(año_actual)]
            if not año_row.empty:
                sus_volumen = año_row.iloc[0]["sus_en_mis_aforo"]
        else:
            # Para concesión individual, estimar usando su participación en el aforo total del año
            if sim_target != "Portafolio completo" and not hoja2_df.empty:
                año_actual = selected_year if year_choice != "Todos" else annual_all["year"].max()
                año_row = hoja2_df[hoja2_df["periodo"] == str(año_actual)]
                if not año_row.empty:
                    total_aforo_anual = raw_df[raw_df["year"] == año_actual]["aforo"].sum()
                    if total_aforo_anual > 0:
                        participacion = base_aforo / total_aforo_anual
                        sus_volumen = año_row.iloc[0]["sus_en_mis_aforo"] * participacion

        # Mostrar métricas base
        c1, c2, c3 = st.columns(3)
        c1.metric("Base aforo", fmt_num(base_aforo))
        c2.metric("Base ingreso", fmt_mxn(base_ingreso))
        c3.metric("Ingreso / cruce", f"{base_rpc:.2f}" if pd.notna(base_rpc) else "-")

        st.markdown("#### Escenario con captura de SUS EN MIS")
        st.write(f"**Volumen SUS EN MIS a capturar:** {fmt_num(sus_volumen)} (estimado para {sim_target})")
        capture_percent = sim_sus_capture / 100.0
        captured_volume = sus_volumen * capture_percent

        # RPC de TeleVía en la base seleccionada
        televia_df = sim_base[sim_base["family"] == "TeleVía"]
        televia_rpc = safe_divide(televia_df["ingreso"].sum(), televia_df["aforo"].sum()) if not televia_df.empty else base_rpc
        st.caption(f"RPC promedio de TeleVía en la base: {televia_rpc:.2f}")

        # Nuevo ingreso por la captura (se asume que los vehículos capturados pagan tarifa de TeleVía)
        new_ingreso_captura = captured_volume * televia_rpc
        # El aforo total no cambia (los vehículos ya estaban contados), pero el ingreso sí
        scenario_ingreso = base_ingreso + new_ingreso_captura
        scenario_aforo = base_aforo  # sin cambio
        scenario_rpc = safe_divide(scenario_ingreso, scenario_aforo)

        # Nuevo market share de TeleVía
        base_televia_aforo = televia_df["aforo"].sum()
        new_televia_aforo = base_televia_aforo + captured_volume
        new_televia_share = new_televia_aforo / scenario_aforo if scenario_aforo > 0 else 0
        base_televia_share = base_televia_aforo / base_aforo if base_aforo > 0 else 0

        delta_ingreso = scenario_ingreso - base_ingreso

        st.markdown("##### Resultados del escenario")
        scen_df = pd.DataFrame([
            {"Métrica": "Base", "Aforo": base_aforo, "Ingreso": base_ingreso, "Ingreso/cruce": base_rpc, "Share TeleVía": base_televia_share},
            {"Métrica": "Escenario", "Aforo": scenario_aforo, "Ingreso": scenario_ingreso, "Ingreso/cruce": scenario_rpc, "Share TeleVía": new_televia_share},
            {"Métrica": "Delta", "Aforo": scenario_aforo - base_aforo, "Ingreso": delta_ingreso, "Ingreso/cruce": scenario_rpc - base_rpc, "Share TeleVía": new_televia_share - base_televia_share},
        ])
        st.dataframe(
            scen_df.style.format({
                "Aforo": "{:,.0f}",
                "Ingreso": "${:,.0f}",
                "Ingreso/cruce": "{:,.2f}",
                "Share TeleVía": "{:.1%}",
            }),
            use_container_width=True,
            height=220,
            hide_index=True,
        )

        fig = go.Figure()
        fig.add_bar(name="Base", x=["Aforo", "Ingreso", "Share TeleVía"], y=[base_aforo, base_ingreso, base_televia_share], marker_color=COLORS["muted"])
        fig.add_bar(name="Escenario", x=["Aforo", "Ingreso", "Share TeleVía"], y=[scenario_aforo, scenario_ingreso, new_televia_share], marker_color=COLORS["televia"])
        fig.update_layout(barmode="group", legend_title_text="", yaxis_title="Valor")
        st.plotly_chart(make_bar_config(fig, height=420), use_container_width=True)

        st.markdown("##### Lectura de palancas")
        st.write(f"- Capturando el **{sim_sus_capture:.1f}%** de los vehículos SUS EN MIS, se incrementa el ingreso en **{fmt_mxn(delta_ingreso)}**.")
        st.write(f"- El share de TeleVía aumenta de **{base_televia_share:.1%}** a **{new_televia_share:.1%}**.")
        st.write("- Ajusta el porcentaje de captura en la barra lateral para explorar distintos escenarios.")

        # También mantener la simulación original (crecimiento de volumen, etc.) - opcional
        st.markdown("#### Simulación adicional (crecimiento de volumen y mejora de RPC)")
        extra_aforo = base_aforo * (sim_volume_growth / 100.0)
        capture_aforo = base_aforo * (sim_share_capture / 100.0)
        scenario_aforo2 = base_aforo + extra_aforo + capture_aforo
        scenario_rpc2 = base_rpc * (1 + sim_rpc_uplift / 100.0) if pd.notna(base_rpc) else np.nan
        scenario_ingreso2 = scenario_aforo2 * scenario_rpc2 if pd.notna(scenario_rpc2) else np.nan
        delta_aforo2 = scenario_aforo2 - base_aforo
        delta_ingreso2 = scenario_ingreso2 - base_ingreso if pd.notna(scenario_ingreso2) else np.nan

        st.dataframe(
            pd.DataFrame([
                {"Métrica": "Base", "Aforo": base_aforo, "Ingreso": base_ingreso, "Ingreso/cruce": base_rpc},
                {"Métrica": "Escenario (crec. + share)", "Aforo": scenario_aforo2, "Ingreso": scenario_ingreso2, "Ingreso/cruce": scenario_rpc2},
                {"Métrica": "Delta", "Aforo": delta_aforo2, "Ingreso": delta_ingreso2, "Ingreso/cruce": (scenario_rpc2 - base_rpc) if pd.notna(scenario_rpc2) else np.nan},
            ]).style.format({
                "Aforo": "{:,.0f}",
                "Ingreso": "${:,.0f}",
                "Ingreso/cruce": "{:,.2f}",
            }),
            use_container_width=True,
            height=180,
            hide_index=True,
        )

# -----------------------------------------------------------------------------
# 4) Action plan
# -----------------------------------------------------------------------------
with tab_plan:
    st.subheader("Plan de acción")

    if opportunities_df.empty:
        st.info("No hay datos suficientes para construir un plan de acción.")
    else:
        top_opp = opportunities_df.head(MAX_TOP_OPPORTUNITIES).copy()
        top_opp["priority_score"] = top_opp["priority_score"].round(2)

        st.markdown("##### Oportunidades prioritarias")
        action_table = top_opp[
            ["concession", "aforo", "rpc", "aforo_yoy", "volatility", "priority_score", "accion", "motivo"]
        ].copy()
        action_table.columns = ["Concesión", "Aforo", "Ingreso/cruce", "YoY aforo", "Volatilidad", "Prioridad", "Acción", "Motivo"]
        st.dataframe(
            action_table.style.format({
                "Aforo": "{:,.0f}",
                "Ingreso/cruce": "{:,.2f}",
                "YoY aforo": "{:.1%}",
                "Volatilidad": "{:.2f}",
                "Prioridad": "{:.2f}",
            }),
            use_container_width=True,
            height=320,
            hide_index=True,
        )

        cols = st.columns([1.1, 0.9])
        with cols[0]:
            st.markdown("##### Priorización")
            plot_df = top_opp.sort_values("priority_score", ascending=True)
            fig = px.bar(
                plot_df,
                x="priority_score",
                y="concession",
                orientation="h",
                color="priority_score",
                color_continuous_scale="RdYlGn_r",
                text_auto=".2f",
                hover_data={
                    "aforo": ":,.0f",
                    "rpc": ":,.2f",
                    "aforo_yoy": ":.1%",
                    "volatility": ":.2f",
                },
            )
            fig.update_layout(xaxis_title="Prioridad", yaxis_title="")
            st.plotly_chart(make_bar_config(fig, height=max(420, 28 * len(top_opp) + 120)), use_container_width=True)

        with cols[1]:
            st.markdown("##### Playbook sugerido")
            st.markdown(
                """
                <div style="background:white;border:1px solid #E2E8F0;border-radius:14px;padding:14px 16px;">
                <ul style="margin:0;padding-left:18px;">
                    <li><b>Recuperación de volumen</b>: donde la concesión cae pero sigue siendo relevante.</li>
                    <li><b>Monetización</b>: cuando el ingreso/cruce está por debajo del benchmark.</li>
                    <li><b>Estabilización</b>: cuando la variabilidad mensual es alta.</li>
                    <li><b>Captura de canal</b>: cuando hay poco peso de PASE o del mix TeleVía.</li>
                    <li><b>Mix premium</b>: cuando el canal de mayor valor puede crecer.</li>
                </ul>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.caption("Ajusta los supuestos con los sliders de simulación en la barra lateral.")

# -----------------------------------------------------------------------------
# 5) Follow-up
# -----------------------------------------------------------------------------
with tab_follow:
    st.subheader("Seguimiento")

    left, right = st.columns([1.1, 0.9])

    with left:
        st.markdown("##### Evolución anual del portafolio")
        annual_plot = annual_all.copy()

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=annual_plot["year"],
                y=annual_plot["aforo"],
                mode="lines+markers",
                name="Aforo",
                line=dict(color=COLORS["televia"], width=4),
            )
        )
        fig.add_trace(
            go.Scatter(
                x=annual_plot["year"],
                y=annual_plot["ingreso"],
                mode="lines+markers",
                name="Ingreso",
                line=dict(color=COLORS["info"], width=4),
                yaxis="y2",
            )
        )
        fig.update_layout(
            xaxis=dict(title="Año"),
            yaxis=dict(title="Aforo", gridcolor=COLORS["grid"]),
            yaxis2=dict(title="Ingreso", overlaying="y", side="right", showgrid=False),
            legend_title_text="",
        )
        st.plotly_chart(make_bar_config(fig, height=420), use_container_width=True)

    with right:
        st.markdown("##### KPI anual")
        annual_display = annual_plot.copy()
        st.dataframe(
            annual_display.rename(columns={
                "year": "Año",
                "aforo": "Aforo",
                "ingreso": "Ingreso",
                "rpc": "Ingreso/cruce",
                "aforo_yoy": "YoY aforo",
                "ingreso_yoy": "YoY ingreso",
            }).style.format({
                "Aforo": "{:,.0f}",
                "Ingreso": "${:,.0f}",
                "Ingreso/cruce": "{:.2f}",
                "YoY aforo": "{:.1%}",
                "YoY ingreso": "{:.1%}",
            }),
            use_container_width=True,
            height=320,
            hide_index=True,
        )

        if len(annual_display) >= 2:
            last = annual_display.iloc[-1]
            prev = annual_display.iloc[-2]
            st.metric("Variación aforo último año", fmt_pct(safe_pct_change(last["aforo"], prev["aforo"])))
            st.metric("Variación ingreso último año", fmt_pct(safe_pct_change(last["ingreso"], prev["ingreso"])))

        st.markdown("##### Hoja2 · captura complementaria")
        if not hoja2_df.empty:
            st.dataframe(hoja2_df, use_container_width=True, height=220, hide_index=True)
        else:
            st.info("Hoja2 no disponible en el archivo.")

    st.markdown("##### Próximo paso")
    st.info(
        "Cuando agregues otro Excel con la misma estructura, la lógica de parseo se conserva. "
        "Solo cambia el archivo en el repositorio y el dashboard recalcula los bloques."
    )

# -----------------------------------------------------------------------------
# 6) Interoperabilidad y Captura (nueva pestaña)
# -----------------------------------------------------------------------------
with tab_interop:
    st.subheader("Análisis de interoperabilidad y potencial de captura")

    if hoja2_df.empty:
        st.warning("No se encontraron datos en la hoja 'Hoja2'. No es posible mostrar el análisis de interoperabilidad.")
    else:
        # Preparar datos anuales
        hoja2_clean = hoja2_df.copy()
        hoja2_clean["año"] = hoja2_clean["periodo"].astype(str)
        # Calcular share de TeleVía en MIS EN MIS usando raw_df
        televia_shares = []
        for idx, row in hoja2_clean.iterrows():
            año = row["periodo"]
            # Intentar convertir a entero
            try:
                año_int = int(año)
            except:
                año_int = None
            if año_int is not None:
                total_aforo_anual = raw_df[raw_df["year"] == año_int]["aforo"].sum()
                televia_aforo = raw_df[(raw_df["year"] == año_int) & (raw_df["family"] == "TeleVía")]["aforo"].sum()
                share = safe_divide(televia_aforo, total_aforo_anual)
            else:
                share = np.nan
            televia_shares.append(share)
        hoja2_clean["share_televia_mis_en_mis"] = televia_shares

        # Potencial de captura: asumimos que actualmente el share de TeleVía en SUS EN MIS es 10% (conservador)
        share_sus_actual = 0.10
        hoja2_clean["potencial_captura_aforo"] = hoja2_clean["sus_en_mis_aforo"] * (1 - share_sus_actual)
        hoja2_clean["potencial_captura_ingreso"] = hoja2_clean["potencial_captura_aforo"] * (hoja2_clean["share_televia_mis_en_mis"] * (hoja2_clean["mis_en_mis_ingreso"] / hoja2_clean["mis_en_mis_aforo"]).fillna(0))

        # Indicador de fuga
        hoja2_clean["fuga_ratio"] = hoja2_clean["mis_en_sus_aforo"] / hoja2_clean["mis_en_mis_aforo"].replace(0, np.nan)

        st.markdown("#### Tabla de flujos anuales")
        display_cols = ["año", "mis_en_mis_aforo", "mis_en_sus_aforo", "sus_en_mis_aforo", "share_televia_mis_en_mis", "fuga_ratio", "potencial_captura_aforo"]
        st.dataframe(
            hoja2_clean[display_cols].rename(columns={
                "año": "Año",
                "mis_en_mis_aforo": "MIS EN MIS (aforo)",
                "mis_en_sus_aforo": "MIS EN SUS (aforo)",
                "sus_en_mis_aforo": "SUS EN MIS (aforo)",
                "share_televia_mis_en_mis": "Share TeleVía en MIS EN MIS",
                "fuga_ratio": "Ratio fuga (MIS EN SUS / MIS EN MIS)",
                "potencial_captura_aforo": "Potencial captura SUS EN MIS (aforo)",
            }).style.format({
                "MIS EN MIS (aforo)": "{:,.0f}",
                "MIS EN SUS (aforo)": "{:,.0f}",
                "SUS EN MIS (aforo)": "{:,.0f}",
                "Share TeleVía en MIS EN MIS": "{:.1%}",
                "Ratio fuga (MIS EN SUS / MIS EN MIS)": "{:.2%}",
                "Potencial captura SUS EN MIS (aforo)": "{:,.0f}",
            }),
            use_container_width=True,
            height=320,
            hide_index=True,
        )

        st.markdown("#### Gráfico de barras apiladas de flujos anuales")
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(x=hoja2_clean["año"], y=hoja2_clean["mis_en_mis_aforo"], name="MIS EN MIS", marker_color=COLORS["televia"]))
        fig_bar.add_trace(go.Bar(x=hoja2_clean["año"], y=hoja2_clean["mis_en_sus_aforo"], name="MIS EN SUS", marker_color=COLORS["warning"]))
        fig_bar.add_trace(go.Bar(x=hoja2_clean["año"], y=hoja2_clean["sus_en_mis_aforo"], name="SUS EN MIS", marker_color=COLORS["info"]))
        fig_bar.update_layout(barmode="stack", xaxis_title="Año", yaxis_title="Aforo", legend_title="Flujo")
        st.plotly_chart(make_bar_config(fig_bar, height=450), use_container_width=True)

        st.markdown("#### Diagrama Sankey de flujos")
        # Selector de año para Sankey
        año_sankey = st.selectbox("Selecciona un año para el diagrama Sankey", hoja2_clean["año"].unique(), key="sankey_year")
        sankey_data = hoja2_clean[hoja2_clean["año"] == año_sankey].iloc[0]
        # Nodos: Origen y destino: MIS y SUS
        labels = ["MIS", "SUS"]
        # Fuentes: 0=MIS, 1=SUS
        source = [0, 0, 1]  # MIS->MIS, MIS->SUS, SUS->MIS
        target = [0, 1, 0]
        values = [
            sankey_data["mis_en_mis_aforo"],
            sankey_data["mis_en_sus_aforo"],
            sankey_data["sus_en_mis_aforo"]
        ]
        fig_sankey = go.Figure(data=[go.Sankey(
            node=dict(
                pad=15,
                thickness=20,
                line=dict(color="black", width=0.5),
                label=labels,
                color=[COLORS["televia"], COLORS["info"]]
            ),
            link=dict(
                source=source,
                target=target,
                value=values,
                color=[COLORS["televia"], COLORS["warning"], COLORS["info"]]
            )
        )])
        fig_sankey.update_layout(title=f"Flujos de tráfico {año_sankey}", height=500)
        st.plotly_chart(fig_sankey, use_container_width=True)

        st.markdown("#### Interpretación")
        st.info(
            "**MIS EN MIS**: Vehículos de la propia concesionaria que usan sus propios carriles.\n\n"
            "**MIS EN SUS**: Vehículos propios que transitan por carriles de otros operadores (fuga).\n\n"
            "**SUS EN MIS**: Vehículos de otros operadores que usan nuestras autopistas (oportunidad de captura).\n\n"
            "**Potencial de captura**: Representa el volumen de SUS EN MIS que aún no usa TeleVía. Convertirlos aumentaría el share y el ingreso."
        )

# -----------------------------------------------------------------------------
# Footer
# -----------------------------------------------------------------------------
st.markdown(
    f"<div style='margin-top:1.2rem; color:{COLORS['muted']}; font-size:0.85rem;'>"
    f"Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M')} · Fuente: {EXCEL_FILE}"
    f"</div>",
    unsafe_allow_html=True,
)
